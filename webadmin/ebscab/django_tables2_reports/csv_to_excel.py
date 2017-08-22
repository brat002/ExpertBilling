# -*- coding: utf-8 -*-
# Copyright (c) 2010 by Yaco Sistemas <pmartin@yaco.es>
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Lesser General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Lesser General Public License for more details.
#
# You should have received a copy of the GNU Lesser General Public License
# along with this programe.  If not, see <http://www.gnu.org/licenses/>.

# Based on http://sujitpal.blogspot.com/2007/02/python-script-to-convert-csv-files-to.html
# Get to
# https://github.com/Yaco-Sistemas/django-autoreports/blob/master/autoreports/csv_to_excel.py

import csv

from openpyxl.workbook import Workbook
from openpyxl.writer.excel import save_virtual_workbook


HAS_PYEXCELERATOR = True


def openExcelSheet():
    """ Opens a reference to an Excel WorkBook and Worksheet objects """
    workbook = Workbook()
    worksheet = workbook.create_sheet(0)
    return workbook, worksheet


def validateOpts(response):
    """ Returns option values specified, or the default if none """
    titlePresent = False
    linesPerFile = -1
    sepChar = ","
    return titlePresent, linesPerFile, sepChar


def closeExcelSheet(response, workbook):
    """ Saves the in-memory WorkBook object into the specified file """
    response.content = save_virtual_workbook(workbook)
    # response.content_type='application/vnd.ms-excel'
    # return HttpResponse(save_virtual_workbook(workbook),
    # content_type='application/vnd.ms-excel')


def convert_to_excel(response):
    titlePresent, linesPerFile, sepChar = validateOpts(response)
    workbook, worksheet = openExcelSheet()
    fno = 0
    lno = 0
    titleCols = []
    content = response.content.split('\n')
    reader = csv.reader(content)
    for row_index, row in enumerate(reader):
        for col_index, col in enumerate(row):
            worksheet.cell(
                row=row_index, column=col_index).value = col.decode('utf-8')
    closeExcelSheet(response, workbook)
